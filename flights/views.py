# flights/views.py
import logging
logger = logging.getLogger(__name__)
from datetime import date, datetime
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.mail import send_mail
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db import transaction
from django.db import IntegrityError
from django.db import connection
from django.db.models import Q, Count, Sum, F, Exists, OuterRef
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.template.loader import render_to_string
from django.http import HttpResponse, JsonResponse
from accounts.forms import CustomUserChangeForm
from accounts.models import CustomUser
from django.views.decorators.http import require_GET
from .forms import (
    BookingForm, FlightCreateForm, FlightEditForm, FlightSearchForm,
    FlightSelectForm, AirportForm, SeatSelectionForm, PaymentConfirmationForm
)
from .models import Airline, Airport, Booking, Flight, Seat, Payment
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import pandas as pd
import pytz
from decimal import Decimal


def home(request):
    today = date.today()
    featured_flights = Flight.objects.filter(
        is_active=True,
        departure_time__gte=date.today()
    ).order_by('departure_time')[:6]
    
    airports = Airport.objects.all()
    
    if request.method == 'POST':
        form = FlightSearchForm(request.POST)
        if form.is_valid():
            origin = form.cleaned_data['origin']
            destination = form.cleaned_data['destination']
            departure_date = form.cleaned_data['departure_date']
            
            flights = Flight.objects.filter(
                departure_airport=origin,
                arrival_airport=destination,
                departure_time__date=departure_date,
                is_active=True,
                available_seats__gt=0
            ).order_by('departure_time')
            
            return render(request, 'flights/flight_list.html', {
                'flights': flights,
                'origin': origin,
                'destination': destination,
                'departure_date': departure_date,
            })
    else:
        form = FlightSearchForm()
    
    return render(request, 'home.html', {
        'today': today,
        'featured_flights': featured_flights,
        'airports': airports,
        'form': form,
    })

def flight_search(request):
    if request.method == 'POST':
        form = FlightSearchForm(request.POST)
        if form.is_valid():
            origin = form.cleaned_data['origin']
            destination = form.cleaned_data['destination']
            departure_date = form.cleaned_data['departure_date']
            
            flights = Flight.objects.filter(
                Q(departure_airport=origin) &
                Q(arrival_airport=destination) &
                Q(departure_time__date=departure_date) &
                Q(is_active=True) &
                Q(available_seats__gt=0)
            ).order_by('departure_time')
            
            return render(request, 'flights/flight_list.html', {
                'flights': flights,
                'origin': origin,
                'destination': destination,
                'departure_date': departure_date,
                'form': form
            })
    else:
        form = FlightSearchForm()
    
    return render(request, 'flights/flight_list.html', {'form': form})

@login_required
def book_flight(request, flight_id):
    flight = get_object_or_404(Flight, id=flight_id, is_active=True)
    seats = Seat.objects.filter(flight=flight).order_by('seat_number')
    
    if not seats.exists():
        messages.error(request, "Este vuelo no tiene asientos configurados.")
        return redirect('flight_search')  # Sin namespace

    if request.method == 'POST':
        form = BookingForm(request.POST, flight=flight)
        selected_seats = request.POST.getlist('seats')
        
        if form.is_valid():
            try:
                with transaction.atomic():
                    passengers = form.cleaned_data['passengers']
                    
                    # Validaciones
                    if flight.available_seats < passengers:
                        messages.error(request, "No hay suficientes asientos disponibles")
                        return redirect('book_flight', flight_id=flight.id)
                    
                    if len(selected_seats) != passengers:
                        messages.error(request, f"Debes seleccionar exactamente {passengers} asientos")
                        return redirect('book_flight', flight_id=flight.id)
                    
                    # Crear reserva
                    booking = Booking.objects.create(
                        user=request.user,
                        flight=flight,
                        passengers=passengers,
                        total_price=flight.price * passengers,
                        status='PENDING'
                    )
                    
                    # Actualizar asientos
                    seats_to_update = seats.filter(seat_number__in=selected_seats, status='AVAILABLE')
                    if seats_to_update.count() != passengers:
                        messages.error(request, "Algunos asientos ya no están disponibles")
                        raise Exception("Asientos no disponibles")
                    
                    seats_to_update.update(status='RESERVED', booking=booking)
                    
                    
                    # Redirigir a la confirmación de pago
                    return redirect('payment_confirmation', booking_id=booking.id)
            
            except Exception as e:
                messages.error(request, f"Error al crear la reserva: {str(e)}")
                logger.error(f"Error en book_flight: {str(e)}", exc_info=True)
        else:
            messages.error(request, "Por favor corrige los errores en el formulario")
    else:
        form = BookingForm(flight=flight)
    
    # Organizar asientos para la vista
    seats_per_row = 6
    seat_rows = []
    all_seats = list(seats)
    
    for i in range(0, len(all_seats), seats_per_row):
        row_seats = all_seats[i:i+seats_per_row]
        seat_rows.append(row_seats)
    
    return render(request, 'flights/booking_form.html', {
        'flight': flight,
        'form': form,
        'seat_rows': seat_rows,
        'selected_seats': request.POST.getlist('seats', [])
    })

@login_required
def payment_confirmation(request, booking_id):
    booking = get_object_or_404(Booking, id=booking_id, user=request.user)
    
    if booking.status == 'CONFIRMED':
        messages.warning(request, 'Esta reserva ya ha sido confirmada')
        return redirect('booking_detail', booking_id=booking.id)
    
    # Obtener o crear el pago
    payment, created = Payment.objects.get_or_create(
        booking=booking,
        defaults={
            'amount': booking.total_price,
            'payment_method': 'QR'
        }
    )
    
    if request.method == 'POST':
        form = PaymentConfirmationForm(request.POST, instance=payment)  # Usa instance correctamente
        if form.is_valid():
            try:
                payment.status = 'COMPLETED'
                payment.save()
                
                # Actualizar estado de la reserva
                booking.status = 'CONFIRMED'
                booking.save()
                
                booking.send_confirmation_email()
                
                # Actualizar asientos
                booking.booked_seats.update(status='OCCUPIED')
                
                messages.success(request, 'Pago completado y reserva confirmada!')
                return redirect('booking_detail', booking_id=booking.id)
            except Exception as e:
                messages.error(request, f"Error al procesar el pago: {str(e)}")
    else:
        form = PaymentConfirmationForm(instance=payment)  # Pasa la instancia correctamente
    
    return render(request, 'flights/payment_confirmation.html', {
        'booking': booking,
        'payment': payment,
        'form': form,
    })

@login_required
def booking_detail(request, booking_id):
    booking = get_object_or_404(
        Booking.objects.select_related('flight', 'user', 'payment'),
        id=booking_id,
        user=request.user
    )
    
    return render(request, 'flights/booking_detail.html', {
        'booking': booking
    })

@login_required
def select_flight(request, flight_id):
    flight = get_object_or_404(Flight, id=flight_id, is_active=True)
    
    if request.method == 'POST':
        form = FlightSelectForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Verificar disponibilidad
                    if flight.available_seats < form.cleaned_data['passengers']:
                        messages.error(request, "No hay suficientes asientos disponibles")
                        return redirect('select_flight', flight_id=flight.id)
                    
                    # Crear booking en estado "seleccionado"
                    booking = Booking.objects.create(
                        user=request.user,
                        flight=flight,
                        passengers=form.cleaned_data['passengers'],
                        total_price=flight.price * form.cleaned_data['passengers'],
                        status='SELECTED'
                    )
                    
                    messages.success(request, 'Vuelo seleccionado. Por favor complete los detalles de la reserva.')
                    return redirect('book_flight', flight_id=flight.id)
            
            except Exception as e:
                messages.error(request, f"Error al seleccionar vuelo: {str(e)}")
    else:
        form = FlightSelectForm()
    
    return render(request, 'flights/select_flight.html', {
        'flight': flight,
        'form': form
    })

@login_required
def confirm_booking(request, booking_id):
    booking = get_object_or_404(Booking, id=booking_id, user=request.user)
    selected_seats = request.session.get('selected_seats', [])
    
    try:
        with transaction.atomic():
            # Obtener y validar asientos
            seats = Seat.objects.filter(
                id__in=selected_seats, 
                status='AVAILABLE',
                flight=booking.flight
            )
            
            if seats.count() != booking.passengers:
                messages.error(request, "Algunos asientos ya no están disponibles")
                return redirect('select_seats', flight_id=booking.flight.id)
            
            # Actualizar asientos y reserva
            seats.update(status='RESERVED', booking=booking)
            booking.seats.set(seats)
            booking.status = 'CONFIRMED'
            booking.save()
            
            messages.success(request, "Reserva confirmada exitosamente!")
            return redirect('booking_detail', booking_id=booking.id)
            
    except Exception as e:
        logger.error(f"Error confirmando reserva: {str(e)}")
        messages.error(request, "Ocurrió un error al confirmar la reserva")
        return redirect('dashboard')

def is_admin(user):
    return user.is_authenticated and user.is_staff

@user_passes_test(is_admin)
def admin_flight_list(request):
    if not request.user.is_staff:
        return redirect('dashboard')
    
    # Filtros
    flight_number = request.GET.get('flight_number')
    origin = request.GET.get('origin')
    destination = request.GET.get('destination')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    airline = request.GET.get('airline')
    status = request.GET.get('status', 'active')  # active/inactive/all
    
    # Consulta base con select_related
    flights = Flight.objects.all().select_related(
        'airline', 'departure_airport', 'arrival_airport'
    )
    
    # Aplicar filtros
    if flight_number:
        flights = flights.filter(flight_number__icontains=flight_number)
    if origin:
        flights = flights.filter(departure_airport__code__icontains=origin)
    if destination:
        flights = flights.filter(arrival_airport__code__icontains=destination)
    if date_from:
        flights = flights.filter(departure_time__date__gte=date_from)
    if date_to:
        flights = flights.filter(departure_time__date__lte=date_to)
    if airline:
        flights = flights.filter(airline__name__icontains=airline)
    if status == 'active':
        flights = flights.filter(is_active=True)
    elif status == 'inactive':
        flights = flights.filter(is_active=False)
    
    flights = flights.order_by('-departure_time')
    
    # Paginación
    paginator = Paginator(flights, 15)
    page = request.GET.get('page')
    
    try:
        flights = paginator.page(page)
    except PageNotAnInteger:
        flights = paginator.page(1)
    except EmptyPage:
        flights = paginator.page(paginator.num_pages)
    
    # Obtener aeropuertos para el dropdown de filtros
    airports = Airport.objects.all().order_by('city')
    airlines = Airline.objects.all().order_by('name')
    
    return render(request, 'flights/admin/flight_list.html', {
        'flights': flights,
        'airports': airports,
        'airlines': airlines,
        'filter_values': {
            'origin': origin or '',
            'destination': destination or '',
            'date_from': date_from or '',
            'date_to': date_to or '',
            'airline': airline or '',
            'status': status,
        }
    })


@user_passes_test(is_admin)
def admin_create_flight(request):
    try:
        # Obtener la aerolínea Velair
        velair = Airline.objects.get(name="Velair")  # Asegúrate que el nombre coincida exactamente
        
        if request.method == 'POST':
            form = FlightCreateForm(request.POST)
            if form.is_valid():
                flight = form.save(commit=False)
                flight.airline = velair  # Asignar aerolínea
                flight.save()
                messages.success(request, 'Vuelo creado exitosamente!')
                return redirect('admin_flight_list')
        else:
            form = FlightCreateForm()
            
        return render(request, 'flights/admin/flight_form.html', {
            'form': form,
            'title': 'Crear nuevo vuelo'
        })
        
    except Airline.DoesNotExist:
        messages.error(request, 'Error: Primero debe crear la aerolínea "Velair"')
        return redirect('admin_dashboard')

@user_passes_test(is_admin)
def admin_edit_flight(request, flight_id):
    flight = get_object_or_404(
        Flight.objects.select_related('airline', 'departure_airport', 'arrival_airport'), 
        id=flight_id
    )

    if flight.departure_time < timezone.now():
        messages.error(request, "No se puede editar un vuelo que ya ha despegado.")
        return redirect('admin_flight_list')
    
    if request.method == 'POST':
        # Manejar eliminación de reserva
        if 'remove_booking' in request.POST:
            booking_id = request.POST.get('remove_booking')
            try:
                booking = Booking.objects.get(id=booking_id, flight=flight)
                
                # Mensaje según estado
                status_msg = {
                    'CONFIRMED': 'Se notificará sobre el reembolso.',
                    'PENDING': 'Se cancelará el proceso de pago.',
                }.get(booking.status, '')
                
                if booking.cancel(notify_user=True):
                    messages.success(request, 'Reserva cancelada correctamente')
                else:
                    messages.warning(request, 'La reserva ya estaba cancelada')
                
                return redirect('admin_edit_flight', flight_id=flight.id)
            except Booking.DoesNotExist:
                messages.error(request, 'La reserva no existe')
            except Exception as e:
                messages.error(request, f'Error al cancelar: {str(e)}')
            return redirect('admin_edit_flight', flight_id=flight.id)
        
        # Resto del código para actualizar el vuelo...
        form = FlightEditForm(request.POST, instance=flight)
        if form.is_valid():
            form.save()
            messages.success(request, 'Vuelo actualizado exitosamente!')
            return redirect('admin_edit_flight', flight_id=flight.id)
    else:
        form = FlightEditForm(instance=flight)
    
    bookings = Booking.objects.filter(flight=flight).select_related('user', 'payment')
    return render(request, 'flights/admin/flight_form.html', {
        'form': form,
        'title': 'Editar Vuelo',
        'flight': flight,
        'bookings': bookings,
    })

@user_passes_test(is_admin)
def admin_delete_flight(request, flight_id):
    flight = get_object_or_404(Flight, id=flight_id)
    
    if request.method == 'POST':
        try:
            flight.delete()
            messages.success(request, 'Vuelo eliminado correctamente.')
            return redirect('admin_flight_list')
        except Exception as e:
            messages.error(request, f'Error al eliminar: {str(e)}')
            return redirect('admin_flight_list')
    
    # Verifica si tiene reservas asociadas
    has_bookings = Booking.objects.filter(flight=flight).exists()
    
    return render(request, 'flights/admin/flight_confirm_delete.html', {
        'flight': flight,
        'has_bookings': has_bookings
    })


@user_passes_test(is_admin)
def admin_export_flights(request):
    try:
        # Replicar misma lógica de filtros
        flights = Flight.objects.all().select_related(
            'airline', 'departure_airport', 'arrival_airport'
        )
        
        # Aplicar filtros (misma lógica que admin_flight_list)
        params = request.GET.dict()
        
        if params.get('flight_number'):
            flights = flights.filter(flight_number__icontains=params['flight_number'])
        if params.get('origin'):
            flights = flights.filter(departure_airport__code__icontains=params['origin'])
        if params.get('destination'):
            flights = flights.filter(arrival_airport__code__icontains=params['destination'])
        if params.get('date_from'):
            flights = flights.filter(departure_time__date__gte=params['date_from'])
        if params.get('date_to'):
            flights = flights.filter(departure_time__date__lte=params['date_to'])
        if params.get('airline'):
            flights = flights.filter(airline__name__icontains=params['airline'])
        if params.get('status', 'active') == 'active':
            flights = flights.filter(is_active=True)
        elif params.get('status') == 'inactive':
            flights = flights.filter(is_active=False)
        
        # Optimizar consulta
        flights = flights.order_by('-departure_time')

        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Vuelos"

        # Configurar cabeceras
        headers = [
            ("Número Vuelo", 20),
            ("Aerolínea", 25),
            ("Origen", 15),
            ("Destino", 15),
            ("Salida", 25),
            ("Llegada", 25),
            ("Asientos Totales", 15),
            ("Disponibles", 15),
            ("Estado", 15),
        ]

        for col_num, (header, width) in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
            ws.cell(row=1, column=col_num, value=header).style = 'Headline 2'

        # Llenar datos
        for row_num, flight in enumerate(flights, 2):
            ws.cell(row=row_num, column=1, value=flight.flight_number)
            ws.cell(row=row_num, column=2, value=str(flight.airline))
            ws.cell(row=row_num, column=3, value=flight.departure_airport.code)
            ws.cell(row=row_num, column=4, value=flight.arrival_airport.code)
            ws.cell(row=row_num, column=5, value=flight.departure_time.astimezone().isoformat())
            ws.cell(row=row_num, column=6, value=flight.arrival_time.astimezone().isoformat())
            ws.cell(row=row_num, column=7, value=flight.total_seats)
            ws.cell(row=row_num, column=8, value=flight.available_seats)
            ws.cell(row=row_num, column=9, value="Activo" if flight.is_active else "Inactivo")

        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename="Vuelos_Exportados_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
            },
        )
        wb.save(response)
        return response

    except Exception as e:
        messages.error(request, f"Error al exportar: {str(e)}")
        return redirect('admin_flight_list')



# @user_passes_test(is_admin)
# @transaction.atomic
# def admin_import_flights(request):
#     if request.method == 'POST':
#         try:
#             file = request.FILES['file']
#             df = pd.read_excel(file, engine='openpyxl')

#             # Resetear secuencia de manera confiable
#             try:
#                 with connection.cursor() as cursor:
#                     cursor.execute(
#                         "SELECT setval(pg_get_serial_sequence('flights_flight', 'id'), "
#                         "coalesce(MAX(id), 1), true) FROM flights_flight;"  # true en lugar de false
#                     )
#                     cursor.execute("SELECT currval(pg_get_serial_sequence('flights_flight', 'id'))")
#                     current_seq = cursor.fetchone()[0]
#                     print(f"Secuencia ajustada a: {current_seq}")
#             except Exception as e:
#                 print(f"Error crítico ajustando secuencia: {str(e)}")
#                 raise

#             # Debug: Mostrar estructura del DataFrame
#             print("\n" + "="*50)
#             print("ESTRUCTURA DEL ARCHIVO EXCEL:")
#             print("Columnas:", df.columns.tolist())
#             print("Primeras 3 filas:")
#             print(df.head(3))
            
#             # Eliminar columna ID si existe
#             if 'id' in df.columns:
#                 print("\n¡Advertencia: Columna 'id' detectada! Eliminándola...")
#                 df = df.drop(columns=['id'])
            
#             # Convertir a lista de diccionarios
#             data = df.to_dict('records')
            
#             success = 0
#             errors = []
            
#             for index, record in enumerate(data):
#                 try:
#                     print("\n" + "-"*50)
#                     print(f"PROCESANDO FILA {index+2}:")
#                     print(record)
                    
#                     with transaction.atomic():
#                         # ===== 1. VALIDAR AEROPUERTOS =====
#                         origen = record['Origen'].strip().upper()
#                         destino = record['Destino'].strip().upper()
                        
#                         print(f"Buscando aeropuerto ORIGEN: {origen}")
#                         origin = Airport.objects.get(code=origen)
#                         print(f"Aeropuerto encontrado: {origin}")
                        
#                         print(f"Buscando aeropuerto DESTINO: {destino}")
#                         destination = Airport.objects.get(code=destino)
#                         print(f"Aeropuerto encontrado: {destination}")
                        
#                         # ===== 2. VALIDAR AEROLÍNEA =====
#                         airline_name = record['Aerolínea'].strip()
#                         print(f"Buscando aerolínea: {airline_name}")
#                         airline = Airline.objects.get(name__iexact=airline_name)
#                         print(f"Aerolínea encontrada: {airline}")
                        
#                         # ===== 3. CONVERTIR FECHAS =====
#                         print("Procesando fechas...")
#                         tz_origin = pytz.timezone(origin.timezone)
#                         departure = tz_origin.localize(
#                             pd.to_datetime(record['Salida']).to_pydatetime()
#                         )
                        
#                         tz_dest = pytz.timezone(destination.timezone)
#                         arrival = tz_dest.localize(
#                             pd.to_datetime(record['Llegada']).to_pydatetime()
#                         )
#                         print(f"Salida (tz): {departure}")
#                         print(f"Llegada (tz): {arrival}")
                        
#                         # ===== 4. VALIDAR PRECIO =====
#                         print(f"Precio crudo: {record['Precio']} ({type(record['Precio'])})")
#                         precio = Decimal(str(record['Precio'])).quantize(Decimal('0.00'))
#                         print(f"Precio convertido: {precio}")
                        
#                         if precio <= 0:
#                             raise ValueError("El precio debe ser mayor a 0")
                        
#                         # ===== 5. CREAR/ACTUALIZAR VUELO =====
#                         flight_number = record['Número Vuelo'].strip().upper()
#                         print(f"Intentando crear/actualizar vuelo {flight_number}...")
                        
#                         # Usar update_or_create para evitar duplicados
#                         obj, created = Flight.objects.update_or_create(
#                             flight_number=flight_number,
#                             defaults={
#                                 'departure_airport': origin,
#                                 'arrival_airport': destination,
#                                 'departure_time': departure,
#                                 'arrival_time': arrival,
#                                 'total_seats': record['Asientos Totales'],
#                                 'available_seats': record['Disponibles'],
#                                 'is_active': record['Estado'].strip().lower() == 'activo',
#                                 'airline': airline,
#                                 'price': precio
#                             }
#                         )
                        
#                         success += 1
#                         print(f"Vuelo {'creado' if created else 'actualizado'} con éxito. ID: {obj.id}")

#                 except Exception as e:
#                     error_msg = f"Fila {index+2}: {str(e)}"
#                     print(f"\n❌ ERROR: {error_msg}")
#                     errors.append(error_msg)
            
#             # Resultados finales
#             print("\n" + "="*50)
#             print(f"RESUMEN DE IMPORTACIÓN:")
#             print(f"Éxitos: {success}")
#             print(f"Errores: {len(errors)}")
            
#             if errors:
#                 messages.warning(request, f"Importación parcial: {success} éxitos, {len(errors)} errores")
#                 print("\nERRORES DETALLADOS:")
#                 for error in errors[:5]:  # Mostrar primeros 5 errores
#                     print(f"- {error}")
#             else:
#                 messages.success(request, f"✅ {success} vuelos importados exitosamente!")
            
#             return redirect('admin_flight_list')       
        
#         except Exception as e:
#             error_msg = f"❌ Error crítico: {str(e)}"
#             print(error_msg)
#             messages.error(request, error_msg)
#             return redirect('admin_flight_list')
    
#     return redirect('admin_flight_list')
@require_GET
def calculate_arrival_time(request):
    dep_id = request.GET.get('dep')
    arr_id = request.GET.get('arr')
    dep_time = request.GET.get('time')
    
    try:
        departure_airport = Airport.objects.get(id=dep_id)
        arrival_airport = Airport.objects.get(id=arr_id)
        departure_time = timezone.make_aware(datetime.fromisoformat(dep_time))
        
        # Calcular usando el modelo
        flight = Flight(
            departure_airport=departure_airport,
            arrival_airport=arrival_airport,
            departure_time=departure_time
        )
        flight.calculate_arrival_time()
        
        return JsonResponse({
            'arrival_time': timezone.localtime(flight.arrival_time).strftime('%Y-%m-%dT%H:%M')
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@user_passes_test(is_admin)
def admin_user_list(request):
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', 'all')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    user_type = request.GET.get('type', 'all')
    
    users = CustomUser.objects.all().order_by('-date_joined')
    
    # Filtros
    if search_query:
        users = users.filter(
            Q(username__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(phone__icontains=search_query)
        )
    
    if status_filter == 'active':
        users = users.filter(is_active=True)
    elif status_filter == 'inactive':
        users = users.filter(is_active=False)
    
    if user_type == 'staff':
        users = users.filter(is_staff=True)
    elif user_type == 'users':
        users = users.filter(is_staff=False)
    
    if date_from:
        users = users.filter(date_joined__date__gte=date_from)
    if date_to:
        users = users.filter(date_joined__date__lte=date_to)
    
    # Anotar con pagos pendientes (usando 'status' en lugar de 'is_paid')
    users = users.annotate(
        has_pending_payments=Exists(
            Payment.objects.filter(
                booking__user=OuterRef('pk'),
                status='PENDING'  # Cambia 'PENDING' por el valor correcto según tu modelo
            )
        )
    )
    
    paginator = Paginator(users, 15)
    page = request.GET.get('page')
    
    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)
    
    return render(request, 'flights/admin/user_list.html', {
        'users': users,
        'search_query': search_query,
        'status_filter': status_filter,
        'date_from': date_from,
        'date_to': date_to,
        'user_type': user_type,
    })

@user_passes_test(is_admin)
def notify_pending_payment(request, user_id):
    if request.method == 'POST':
        user = get_object_or_404(CustomUser, id=user_id)
        pending_payments = Payment.objects.filter(booking__user=user, is_paid=False)
        
        if pending_payments.exists():
            # Enviar correo electrónico
            subject = 'Recordatorio de pago pendiente'
            message = render_to_string('flights/emails/payment_reminder.txt', {
                'user': user,
                'payments': pending_payments
            })
            send_mail(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL,
                [user.email],
                fail_silently=False,
            )
            return JsonResponse({'status': 'success'})
        
        return JsonResponse({'status': 'error', 'message': 'No hay pagos pendientes'}, status=400)
    
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)

@user_passes_test(is_admin)
def admin_user_detail(request, user_id):
    user = get_object_or_404(CustomUser, id=user_id)
    bookings = Booking.objects.filter(user=user).select_related('flight', 'payment')
    
    return render(request, 'flights/admin/user_detail.html', {
        'user': user,
        'bookings': bookings
    })

@user_passes_test(is_admin)
def admin_user_edit(request, user_id):
    user = get_object_or_404(CustomUser, id=user_id)
    
    if request.method == 'POST':
        form = CustomUserChangeForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Usuario actualizado correctamente')
            return redirect('admin_user_detail', user_id=user.id)
    else:
        form = CustomUserChangeForm(instance=user)
    
    return render(request, 'flights/admin/user_edit.html', {
        'form': form,
        'user': user
    })

@user_passes_test(is_admin)
def admin_user_delete(request, user_id):
    user = get_object_or_404(CustomUser, id=user_id)
    
    if request.method == 'POST':
        user.delete()
        messages.success(request, 'Usuario eliminado correctamente')
        return redirect('admin_user_list')
    
    return render(request, 'flights/admin/user_confirm_delete.html', {
        'user': user
    })

@user_passes_test(is_admin)
def admin_airport_list(request):
    airports = Airport.objects.all().order_by('city')
    return render(request, 'flights/admin/airport_list.html', {
        'airports': airports
    })

@user_passes_test(is_admin)
def admin_airport_create(request):
    if request.method == 'POST':
        form = AirportForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Aeropuerto creado exitosamente!')
            return redirect('admin_airport_list')
    else:
        form = AirportForm()
    
    return render(request, 'flights/admin/airport_form.html', {
        'form': form,
        'title': 'Crear nuevo aeropuerto'
    })

@user_passes_test(is_admin)
def admin_airport_edit(request, airport_id):
    airport = get_object_or_404(Airport, id=airport_id)
    
    if request.method == 'POST':
        form = AirportForm(request.POST, instance=airport)
        if form.is_valid():
            form.save()
            messages.success(request, 'Aeropuerto actualizado exitosamente!')
            return redirect('admin_airport_list')
    else:
        form = AirportForm(instance=airport)
    
    return render(request, 'flights/admin/airport_form.html', {
        'form': form,
        'title': 'Editar aeropuerto',
        'airport': airport
    })

@user_passes_test(is_admin)
def admin_airport_delete(request, airport_id):
    airport = get_object_or_404(Airport, id=airport_id)
    
    if request.method == 'POST':
        airport.delete()
        messages.success(request, 'Aeropuerto eliminado correctamente.')
        return redirect('admin_airport_list')
    
    return render(request, 'flights/admin/airport_confirm_delete.html', {
        'airport': airport
    })

@user_passes_test(is_admin)
def admin_payment_list(request):
    payments = Payment.objects.select_related('booking', 'booking__user', 'booking__flight').order_by('-payment_date')
    
    # Filtros
    status = request.GET.get('status', 'all')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    method = request.GET.get('method', 'all')
    
    if status != 'all':
        payments = payments.filter(status=status)
    
    if date_from:
        payments = payments.filter(payment_date__date__gte=date_from)
    if date_to:
        payments = payments.filter(payment_date__date__lte=date_to)
    
    if method != 'all':
        payments = payments.filter(payment_method=method)
    
    paginator = Paginator(payments, 15)
    page = request.GET.get('page')
    
    try:
        payments = paginator.page(page)
    except PageNotAnInteger:
        payments = paginator.page(1)
    except EmptyPage:
        payments = paginator.page(paginator.num_pages)
    
    return render(request, 'flights/admin/payment_list.html', {
        'payments': payments,
        'status': status,
        'date_from': date_from,
        'date_to': date_to,
        'method': method,
    })

@user_passes_test(is_admin)
def admin_payment_detail(request, payment_id):
    payment = get_object_or_404(
        Payment.objects.select_related('booking', 'booking__user', 'booking__flight'),
        id=payment_id
    )
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        try:
            if action == 'complete':
                payment.mark_as_completed()
                messages.success(request, 'Pago marcado como completado')
            elif action == 'fail':
                payment.mark_as_failed()
                messages.success(request, 'Pago marcado como fallido')
            elif action == 'refund':
                payment.process_refund()
                messages.success(request, 'Pago reembolsado')
            
            return redirect('admin_payment_detail', payment_id=payment.id)
        
        except Exception as e:
            messages.error(request, f'Error al procesar la acción: {str(e)}')
    
    return render(request, 'flights/admin/payment_detail.html', {
        'payment': payment
    })

@login_required
def cancel_booking(request, booking_id):
    booking = get_object_or_404(Booking, id=booking_id, user=request.user)
    if request.method == 'POST':
        if booking.status in ['PENDING', 'CONFIRMED']:
            booking.cancel(notify_user=True)
            messages.success(request, 'La reserva fue cancelada correctamente.')
        else:
            messages.warning(request, 'Esta reserva ya está cancelada.')
    return redirect('user_bookings')
